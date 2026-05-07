"""Build Product Costing & Vendor Sourcing Excel workbook per spec.

Project: https://github.com/oyebiyisunday/product-costing-vendor-sourcing
Maintainer: @oyebiyisunday
"""
from __future__ import annotations

import re
import shutil
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

_REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = _REPO_ROOT / "workbook" / "Product_Costing_Vendor_Sourcing.xlsx"

VENDOR_HEADERS = ["Part_ID", "Part_Name", "Unit_Price"]

VENDOR_A_ROWS = [
    ["P001", "Stepper motor NEMA 17", 24.5],
    ["P002", "Main PCB rev C", 18.0],
    ["P003", "M3x12 screw (100 pk)", 4.2],
    ["P004", "Aluminum bracket", 6.75],
    ["P005", "Power cable 1m", 3.1],
]

VENDOR_B_ROWS = [
    ["P001", "Stepper motor NEMA 17", 22.9],
    ["P002", "Main PCB rev C", 19.5],
    ["P006", "Timing belt GT2 1m", 7.4],
    ["P007", "Idler pulley", 2.15],
    ["P008", "Rubber feet (set)", 1.8],
]

VENDOR_C_ROWS = [
    ["P001", "Stepper motor NEMA 17", 26.0],
    ["P003", "M3x12 screw (100 pk)", 3.95],
    ["P009", "USB-C breakout", 5.5],
    ["P010", "Heat-set inserts M3 (50)", 8.25],
    ["P011", "Acrylic panel 3mm", 12.0],
]

# Sample BOM lines (Part_ID must exist on chosen vendor)
BOM_SAMPLES = [
    ("P001", "Vendor_A", 2),
    ("P002", "Vendor_B", 1),
    ("P003", "Vendor_C", 4),
    ("P006", "Vendor_B", 1),
]


def _style_header(ws, max_col: int) -> None:
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(bold=True, color="FFFFFF")
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _add_vendor_sheet(wb: Workbook, name: str, rows: list, table_name: str) -> None:
    ws = wb.create_sheet(name)
    for col, h in enumerate(VENDOR_HEADERS, start=1):
        ws.cell(row=1, column=col, value=h)
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    last_row = 1 + len(rows)
    ref = f"A1:{get_column_letter(len(VENDOR_HEADERS))}{last_row}"
    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)
    _style_header(ws, len(VENDOR_HEADERS))
    for col in range(1, len(VENDOR_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22 if col == 2 else 14
    ws.freeze_panes = "A2"
    for r in range(2, last_row + 1):
        ws.cell(row=r, column=3).number_format = "$#,##0.00"


def _part_name_formula(row: int) -> str:
    return (
        f'=IF($A{row}="","",'
        f'IFNA(SWITCH($C{row},'
        f'"Vendor_A",XLOOKUP($A{row},Vendor_A!$A:$A,Vendor_A!$B:$B),'
        f'"Vendor_B",XLOOKUP($A{row},Vendor_B!$A:$A,Vendor_B!$B:$B),'
        f'"Vendor_C",XLOOKUP($A{row},Vendor_C!$A:$A,Vendor_C!$B:$B)),'
        '"-- not on vendor --"))'
    )


def _unit_price_formula(row: int) -> str:
    return (
        f'=IF($A{row}="","",'
        f'IFNA(SWITCH($C{row},'
        f'"Vendor_A",XLOOKUP($A{row},Vendor_A!$A:$A,Vendor_A!$C:$C),'
        f'"Vendor_B",XLOOKUP($A{row},Vendor_B!$A:$A,Vendor_B!$C:$C),'
        f'"Vendor_C",XLOOKUP($A{row},Vendor_C!$A:$A,Vendor_C!$C:$C)),'
        '""))'
    )


def _add_product_bom(wb: Workbook) -> None:
    ws = wb.create_sheet("Product_BOM")
    headers = [
        "Part_ID",
        "Part_Name",
        "Vendor",
        "Qty_Per_Product",
        "Total_Qty",
        "Unit_Price",
        "Total_Cost",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    # Parameters (outside BOM table)
    ws["I1"] = "Units_to_build"
    ws["I2"] = 1
    ws["I1"].font = Font(bold=True)
    ws["I2"].number_format = "#,##0"

    table_last_row = 40
    ref = f"A1:{get_column_letter(len(headers))}{table_last_row}"
    tab = Table(displayName="tblBOM", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    _style_header(ws, len(headers))

    # Use A1 references here because Excel repairs openpyxl-generated
    # structured references in some desktop builds.
    for r in range(2, table_last_row + 1):
        ws.cell(row=r, column=2, value=_part_name_formula(r))
        ws.cell(row=r, column=5, value=f'=IF($A{r}="","",$D{r}*$I$2)')
        ws.cell(row=r, column=6, value=_unit_price_formula(r))
        ws.cell(
            row=r,
            column=7,
            value=f'=IF($A{r}="","",IF($F{r}="","",$E{r}*$F{r}))',
        )

    # Sample lines
    for i, (pid, vendor, qty) in enumerate(BOM_SAMPLES, start=2):
        ws.cell(row=i, column=1, value=pid)
        ws.cell(row=i, column=3, value=vendor)
        ws.cell(row=i, column=4, value=qty)

    dv = DataValidation(
        type="list",
        formula1='"Vendor_A,Vendor_B,Vendor_C"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid vendor",
        error="Choose Vendor_A, Vendor_B, or Vendor_C.",
    )
    ws.add_data_validation(dv)
    dv.add(f"C2:C{table_last_row}")

    widths = [12, 28, 12, 18, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions["I"].width = 16

    for r in range(2, table_last_row + 1):
        ws.cell(row=r, column=6).number_format = "$#,##0.00"
        ws.cell(row=r, column=7).number_format = "$#,##0.00"

    ws.freeze_panes = "A2"

    note = (
        "Set Units_to_build in I2. Add BOM lines inside the table only. "
        "Pick vendor in column C; unit price pulls from that vendor sheet by Part_ID."
    )
    ws["K1"] = note
    ws["K1"].font = Font(italic=True, color="666666")
    ws.merge_cells("K1:N3")
    ws["K1"].alignment = Alignment(wrap_text=True, vertical="top")


_EMPTY_V = re.compile(rb"<v\s*/>|<v></v>")


def _strip_empty_cached_values(xlsx_path: Path) -> None:
    """Remove empty cached-value elements from formula cells.

    openpyxl writes ``<f>...</f><v></v>`` for formula cells with no precomputed
    result. Excel 365/2021 treats the empty numeric ``<v>`` as invalid and shows
    a "We found a problem with some content" recovery dialog on open. Stripping
    those empty elements lets Excel calculate values lazily on first open, which
    is the desired behavior here (no Excel engine is available at build time).
    """
    tmp = xlsx_path.with_suffix(xlsx_path.suffix + ".tmp")
    with zipfile.ZipFile(xlsx_path) as zin, zipfile.ZipFile(
        tmp, "w", zipfile.ZIP_DEFLATED
    ) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet") and item.filename.endswith(
                ".xml"
            ):
                data = _EMPTY_V.sub(b"", data)
            zout.writestr(item, data)
    shutil.move(str(tmp), str(xlsx_path))


def main() -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    _add_vendor_sheet(wb, "Vendor_A", VENDOR_A_ROWS, "tblVendorA")
    _add_vendor_sheet(wb, "Vendor_B", VENDOR_B_ROWS, "tblVendorB")
    _add_vendor_sheet(wb, "Vendor_C", VENDOR_C_ROWS, "tblVendorC")
    _add_product_bom(wb)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    _strip_empty_cached_values(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
