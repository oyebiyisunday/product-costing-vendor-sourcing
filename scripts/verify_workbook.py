"""Assert committed workbook matches build_vendor_workbook and BOM samples are valid."""
from __future__ import annotations

import sys
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

_REPO = Path(__file__).resolve().parent.parent
_SCRIPTS = Path(__file__).resolve().parent
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))

import build_vendor_workbook as b


def _vendor_part_ids() -> dict[str, set[str]]:
    return {
        "Vendor_A": {r[0] for r in b.VENDOR_A_ROWS},
        "Vendor_B": {r[0] for r in b.VENDOR_B_ROWS},
        "Vendor_C": {r[0] for r in b.VENDOR_C_ROWS},
    }


def _validate_bom_samples() -> list[str]:
    errs: list[str] = []
    by_vendor = _vendor_part_ids()
    for pid, vendor, _qty in b.BOM_SAMPLES:
        ids = by_vendor.get(vendor)
        if ids is None:
            errs.append(f"Unknown vendor in BOM_SAMPLES: {vendor}")
            continue
        if pid not in ids:
            errs.append(f"BOM sample {pid!r} not on {vendor} table")
    return errs


def _compare_workbooks(wb_disk, wb_fresh) -> list[str]:
    errs: list[str] = []
    if wb_disk.sheetnames != wb_fresh.sheetnames:
        errs.append(f"Sheets: disk={wb_disk.sheetnames} fresh={wb_fresh.sheetnames}")
        return errs
    for sn in wb_disk.sheetnames:
        d, f = wb_disk[sn], wb_fresh[sn]
        if d.max_row != f.max_row or d.max_column != f.max_column:
            errs.append(
                f"{sn}: size disk {d.max_row}x{d.max_column} vs fresh {f.max_row}x{f.max_column}"
            )
            continue
        for r in range(1, d.max_row + 1):
            for c in range(1, d.max_column + 1):
                v1, v2 = d.cell(r, c).value, f.cell(r, c).value
                if v1 != v2:
                    ref = f"{sn}!{get_column_letter(c)}{r}"
                    errs.append(f"{ref}: {v1!r} != {v2!r}")
                    if len(errs) >= 30:
                        return errs
    return errs


_MODERN_ONLY_FUNCS = ("XLOOKUP", "SWITCH", "IFNA", "LET", "LAMBDA")


def _validate_excel_formula_compat(path: Path, wb) -> list[str]:
    errs: list[str] = []
    ws = wb["Product_BOM"]
    for row in range(2, 41):
        for col in (2, 5, 6, 7):
            ref = f"Product_BOM!{get_column_letter(col)}{row}"
            formula = ws.cell(row, col).value
            if not isinstance(formula, str) or not formula.startswith("="):
                errs.append(f"{ref}: missing formula")
                continue
            if "[@" in formula or "#REF!" in formula:
                errs.append(f"{ref}: Excel-fragile formula {formula!r}")
            for fn in _MODERN_ONLY_FUNCS:
                if fn in formula:
                    errs.append(
                        f"{ref}: uses {fn} which is not in older Excel "
                        f"(use VLOOKUP/IFERROR/IF instead)"
                    )

    with zipfile.ZipFile(path) as zf:
        for name in zf.namelist():
            if not name.startswith("xl/worksheets/sheet") or not name.endswith(".xml"):
                continue
            data = zf.read(name)
            if b"<v></v>" in data or b"<v/>" in data:
                errs.append(f"{name}: contains empty cached formula value")
    return errs


def main() -> None:
    sample_errs = _validate_bom_samples()
    if sample_errs:
        print("FAIL: BOM sample / vendor data consistency")
        for e in sample_errs:
            print(" ", e)
        raise SystemExit(1)

    path = _REPO / "workbook" / "Product_Costing_Vendor_Sourcing.xlsx"
    if not path.is_file():
        print("FAIL: missing", path)
        raise SystemExit(1)

    wb_disk = load_workbook(path, data_only=False)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        out = Path(tf.name)
    try:
        b.OUT_PATH = out
        b.main()
        wb_fresh = load_workbook(out, data_only=False)
    finally:
        out.unlink(missing_ok=True)

    diff = _compare_workbooks(wb_disk, wb_fresh)
    if diff:
        print("FAIL: committed workbook does not match generator (re-run build script).")
        for e in diff:
            print(" ", e)
        raise SystemExit(1)

    compat = _validate_excel_formula_compat(path, wb_disk)
    if compat:
        print("FAIL: workbook contains formulas/metadata that Excel may repair.")
        for e in compat[:30]:
            print(" ", e)
        raise SystemExit(1)

    print("OK: workbook matches generator; BOM samples resolve on vendor tables.")


if __name__ == "__main__":
    main()
